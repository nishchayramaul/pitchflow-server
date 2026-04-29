"""
Exports creator leads to a styled .xlsx file.

Design tokens (matched to PitchFlow design system):
  Header bg   : #1F1A15  (warm dark)
  Header text : #FFFFFF
  Accent      : #C8603A  (terracotta)
  Row alt     : #FAF7F2  (warm off-white)
  Row base    : #FFFFFF
  Status fills: per-status palette matching pf-status tokens
"""

import io
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── palette ──────────────────────────────────────────────────────────────────

_COL_HEADER_BG   = "1F1A15"
_COL_HEADER_FG   = "FFFFFF"
_COL_ACCENT      = "C8603A"
_COL_ROW_ALT     = "FAF7F2"
_COL_ROW_BASE    = "FFFFFF"
_COL_BORDER      = "E8E0D8"

_STATUS_COLORS: dict[str, tuple[str, str]] = {
    "pending":     ("FFF3C4", "92400E"),   # amber
    "negotiating": ("DBEAFE", "1E40AF"),   # blue
    "completed":   ("DCFCE7", "166534"),   # green
    "rejected":    ("FEE2E2", "991B1B"),   # red
}

_CURRENCY_SYMBOL: dict[str, str] = {
    "USD": "$", "EUR": "€", "GBP": "£", "INR": "₹", "JPY": "¥",
    "CAD": "CA$", "AUD": "A$", "CHF": "Fr", "CNY": "¥", "SGD": "S$",
    "AED": "د.إ", "BRL": "R$", "MXN": "MX$", "KRW": "₩", "HKD": "HK$",
    "SEK": "kr", "NOK": "kr", "NZD": "NZ$", "ZAR": "R", "SAR": "﷼",
}

# ── helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    side = Side(style="thin", color=_COL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _fmt_budget(value: Any, symbol: str) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{symbol}{float(value):,.0f}"
    except (TypeError, ValueError):
        return str(value)


def _fmt_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
        return dt.strftime("%d %b %Y, %H:%M UTC")
    return str(value)


def _fmt_custom(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


# ── main export ───────────────────────────────────────────────────────────────

def build_leads_excel(
    leads: list[dict[str, Any]],
    form_schema: Optional[list[dict[str, Any]]],
    currency: str = "USD",
    creator_email: str = "",
) -> bytes:
    currency_symbol = _CURRENCY_SYMBOL.get(currency, "$")

    # Determine custom columns from schema (exclude locked fields)
    locked = {"brand_name", "brand_email", "budget"}
    custom_cols: list[tuple[str, str]] = []  # (field_id, label)

    if form_schema:
        for field in form_schema:
            fid = field.get("id", "")
            if fid and fid not in locked:
                custom_cols.append((fid, field.get("label", fid.replace("_", " ").title())))
    else:
        # Fall back: collect all custom_response keys across rows
        seen: dict[str, str] = {}
        for lead in leads:
            for k in (lead.get("custom_responses") or {}):
                if k not in locked and k not in seen:
                    seen[k] = k.replace("_", " ").title()
        custom_cols = list(seen.items())

    # Column definitions: (header_label, extractor_fn)
    columns: list[tuple[str, Any]] = [
        ("#",            lambda i, _l, _c: i),
        ("Brand",        lambda _i, l, _c: l.get("brand_name") or ""),
        ("Email",        lambda _i, l, _c: l.get("brand_email") or ""),
        ("Budget",       lambda _i, l, c: _fmt_budget(l.get("budget"), c)),
        ("Status",       lambda _i, l, _c: (l.get("status") or "").capitalize()),
        ("Submitted",    lambda _i, l, _c: _fmt_date(l.get("created_at"))),
    ]
    for fid, label in custom_cols:
        def _make_extractor(_fid: str):
            return lambda _i, l, _c: _fmt_custom((l.get("custom_responses") or {}).get(_fid))
        columns.append((label, _make_extractor(fid)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── meta sheet ────────────────────────────────────────────────────────────
    meta = wb.create_sheet("Export Info")
    meta.sheet_properties.tabColor = _COL_ACCENT
    meta_rows = [
        ("PitchFlow — Lead Export",),
        ("",),
        ("Exported by", creator_email),
        ("Exported at", datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")),
        ("Total records", len(leads)),
        ("Currency", currency),
    ]
    for r, row in enumerate(meta_rows, 1):
        for c, val in enumerate(row, 1):
            cell = meta.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=13, color=_COL_ACCENT)
            elif c == 1:
                cell.font = Font(bold=True, color=_COL_HEADER_BG)
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 40

    # ── leads sheet ───────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = _COL_ACCENT

    # Freeze pane below header
    ws.freeze_panes = "A2"

    header_font   = Font(bold=True, color=_COL_HEADER_FG, size=10, name="Calibri")
    header_fill   = _fill(_COL_HEADER_BG)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font     = Font(size=10, name="Calibri")
    thin_border   = _thin_border()

    # Write headers
    for col_idx, (label, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # Write data rows
    for row_idx, lead in enumerate(leads, 2):
        is_alt = row_idx % 2 == 0
        row_fill = _fill(_COL_ROW_ALT if is_alt else _COL_ROW_BASE)

        for col_idx, (label, extractor) in enumerate(columns, 1):
            value = extractor(row_idx - 1, lead, currency_symbol)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border

            # Status column: colored pill-style fill
            if label == "Status":
                status_key = (lead.get("status") or "").lower()
                colors = _STATUS_COLORS.get(status_key, (_COL_ROW_ALT, _COL_HEADER_BG))
                cell.fill = _fill(colors[0])
                cell.font = Font(size=10, bold=True, color=colors[1], name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif label == "#":
                cell.fill = _fill(_COL_ROW_ALT if is_alt else _COL_ROW_BASE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=10, color="999090", name="Calibri")
            elif label == "Budget":
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        ws.row_dimensions[row_idx].height = 22

    # ── column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "#":         5,
        "Brand":     22,
        "Email":     30,
        "Budget":    14,
        "Status":    14,
        "Submitted": 24,
    }
    for col_idx, (label, _) in enumerate(columns, 1):
        width = col_widths.get(label, max(len(label) + 4, 18))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── footer accent row ─────────────────────────────────────────────────────
    footer_row = len(leads) + 2
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=footer_row, column=col_idx, value="")
        cell.fill = _fill(_COL_ACCENT)
    ws.row_dimensions[footer_row].height = 4

    # ── serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
